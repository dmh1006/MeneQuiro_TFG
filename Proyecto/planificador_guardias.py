from __future__ import annotations

from collections import defaultdict
from dataclasses import dataclass
from datetime import date, timedelta
from itertools import combinations
from typing import Dict, List, Set, Tuple

import pandas as pd
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ============================================================
# CONFIGURACIÓN
# ============================================================

ANIO = 2026

# Lista de personas del equipo
PERSONAS = [
    "Miembro_01", "Miembro_02", "Miembro_03", "Miembro_04", "Miembro_05",
    "Miembro_06", "Miembro_07", "Miembro_08", "Miembro_09", "Miembro_10",
    "Miembro_11", "Miembro_12", "Miembro_13", "Miembro_14", "Miembro_15",
]

# Festivos a tener en cuenta (ejemplo).
FESTIVOS = {
    date(2026, 1, 1),
    date(2026, 1, 6),
    date(2026, 4, 3),   # Viernes Santo (ejemplo)
    date(2026, 5, 1),
    date(2026, 10, 12),
    date(2026, 11, 1),
    date(2026, 12, 6),
    date(2026, 12, 8),
    date(2026, 12, 25),
}

# Indisponibilidades opcionales por persona
# Ejemplo:
# "Miembro_01": {date(2026, 1, 2), date(2026, 1, 3)}
INDISPONIBILIDADES: Dict[str, Set[date]] = {
    persona: set() for persona in PERSONAS
}

DIAS_ES = {
    0: "Lunes",
    1: "Martes",
    2: "Miércoles",
    3: "Jueves",
    4: "Viernes",
    5: "Sábado",
    6: "Domingo",
}

# ============================================================
# ESTRUCTURAS
# ============================================================

@dataclass
class EstadisticasPersona:
    total: int = 0
    laborables: int = 0
    viernes: int = 0
    sabados: int = 0
    domingos: int = 0
    festivos: int = 0

    def puntuacion_total(self) -> Tuple[int, int, int, int, int, int]:
        """
        Cuanto menor sea esta tupla, más prioritaria es la persona para asignar.
        """
        return (
            self.total,
            self.festivos,
            self.sabados + self.domingos,
            self.laborables,
            self.viernes,
            self.sabados,
        )


# ============================================================
# FUNCIONES DE CALENDARIO
# ============================================================

def generar_fechas_planificables(anio: int) -> List[date]:
    """
    Genera todas las fechas del año excepto julio, agosto y septiembre.
    """
    inicio = date(anio, 1, 1)
    fin = date(anio, 12, 31)

    fechas = []
    actual = inicio
    while actual <= fin:
        if actual.month not in (7, 8, 9):
            fechas.append(actual)
        actual += timedelta(days=1)
    return fechas


def es_festivo(dia: date, festivos: Set[date]) -> bool:
    return dia in festivos


def tipo_dia(dia: date, festivos: Set[date]) -> str:
    """
    Clasificación principal del día.
    """
    if dia in festivos:
        return "festivo"

    wd = dia.weekday()  # lunes=0 ... domingo=6
    if wd == 4:
        return "viernes"
    if wd == 5:
        return "sabado"
    if wd == 6:
        return "domingo"
    return "laborable"


def siguiente_lunes(dia: date) -> date | None:
    """
    Devuelve el lunes posterior a un sábado.
    """
    if dia.weekday() != 5:
        return None
    return dia + timedelta(days=2)


def siguiente_domingo(dia: date) -> date | None:
    """
    Devuelve el domingo posterior a un viernes.
    """
    if dia.weekday() != 4:
        return None
    return dia + timedelta(days=2)


# ============================================================
# LÓGICA DE REPARTO
# ============================================================

def inicializar_estadisticas(personas: List[str]) -> Dict[str, EstadisticasPersona]:
    return {p: EstadisticasPersona() for p in personas}


def crear_bloqueos_por_sabado(
    asignaciones: Dict[date, List[str]]
) -> Dict[date, Set[str]]:
    """
    Si alguien trabaja un sábado, bloqueamos su asignación el lunes siguiente.
    """
    bloqueos = defaultdict(set)

    for dia, personas_dia in asignaciones.items():
        if dia.weekday() == 5:  # sábado
            lunes = dia + timedelta(days=2)
            for persona in personas_dia:
                bloqueos[lunes].add(persona)

    return bloqueos


def actualizar_estadisticas(
    stats: Dict[str, EstadisticasPersona],
    persona: str,
    dia: date,
    festivos: Set[date],
) -> None:
    stats[persona].total += 1

    if es_festivo(dia, festivos):
        stats[persona].festivos += 1

    wd = dia.weekday()
    if wd == 4:
        stats[persona].viernes += 1
    elif wd == 5:
        stats[persona].sabados += 1
    elif wd == 6:
        stats[persona].domingos += 1
    else:
        if not es_festivo(dia, festivos):
            stats[persona].laborables += 1


def calcular_objetivos_teoricos(
    fechas: List[date],
    num_personas: int,
    festivos: Set[date],
    plazas_por_dia: int = 2,
) -> Dict[str, float]:
    """
    Calcula medias teóricas por categoría.
    """
    n_total = len(fechas) * plazas_por_dia
    n_laborables = sum(1 for d in fechas if tipo_dia(d, festivos) == "laborable") * plazas_por_dia
    n_viernes = sum(1 for d in fechas if tipo_dia(d, festivos) == "viernes") * plazas_por_dia
    n_sabados = sum(1 for d in fechas if tipo_dia(d, festivos) == "sabado") * plazas_por_dia
    n_domingos = sum(1 for d in fechas if tipo_dia(d, festivos) == "domingo") * plazas_por_dia
    n_festivos = sum(1 for d in fechas if tipo_dia(d, festivos) == "festivo") * plazas_por_dia

    return {
        "total": n_total / num_personas,
        "laborables": n_laborables / num_personas,
        "viernes": n_viernes / num_personas,
        "sabados": n_sabados / num_personas,
        "domingos": n_domingos / num_personas,
        "festivos": n_festivos / num_personas,
    }


def penalizacion_persona(
    persona: str,
    dia: date,
    stats: Dict[str, EstadisticasPersona],
    objetivos: Dict[str, float],
    asignaciones: Dict[date, List[str]],
    bloqueos_lunes: Dict[date, Set[str]],
    festivos: Set[date],
    indisponibilidades: Dict[str, Set[date]],
) -> float:
    """
    Cuanto menor sea la penalización, mejor candidato es.
    """

    # No disponible ese día
    if dia in indisponibilidades.get(persona, set()):
        return float("inf")

    # Si trabajó sábado, libra lunes
    if persona in bloqueos_lunes.get(dia, set()):
        return float("inf")

    # Evitar repetir en el mismo día
    if persona in asignaciones.get(dia, []):
        return float("inf")

    # Evitar guardias en días consecutivos
    dia_anterior = dia - timedelta(days=1)
    dia_siguiente = dia + timedelta(days=1)

    if persona in asignaciones.get(dia_anterior, []):
        return float("inf")
    if persona in asignaciones.get(dia_siguiente, []):
        return float("inf")

    st = stats[persona]
    t = tipo_dia(dia, festivos)

    pen = 0.0

    # Equilibrio global
    pen += (st.total - objetivos["total"]) * 3

    # Equilibrio por categoría
    if t == "laborable":
        pen += (st.laborables - objetivos["laborables"]) * 4
    elif t == "viernes":
        pen += (st.viernes - objetivos["viernes"]) * 5
    elif t == "sabado":
        pen += (st.sabados - objetivos["sabados"]) * 6
    elif t == "domingo":
        pen += (st.domingos - objetivos["domingos"]) * 6
    elif t == "festivo":
        pen += (st.festivos - objetivos["festivos"]) * 7

    # Penalización ligera por acumulación de fines de semana/festivos
    pen += (st.sabados + st.domingos + st.festivos) * 0.8

    # Si el día es lunes, priorizar que no venga de carga alta reciente
    if dia.weekday() == 0:
        pen += st.total * 0.2

    return pen


def elegir_mejor_pareja(
    candidatos: List[str],
    dia: date,
    stats: Dict[str, EstadisticasPersona],
    objetivos: Dict[str, float],
    asignaciones: Dict[date, List[str]],
    bloqueos_lunes: Dict[date, Set[str]],
    festivos: Set[date],
    indisponibilidades: Dict[str, Set[date]],
) -> Tuple[str, str] | None:
    """
    Elige la mejor pareja de personas para un día concreto.
    """
    mejor_pareja = None
    mejor_pen = float("inf")

    for p1, p2 in combinations(candidatos, 2):
        pen1 = penalizacion_persona(
            p1, dia, stats, objetivos, asignaciones, bloqueos_lunes, festivos, indisponibilidades
        )
        pen2 = penalizacion_persona(
            p2, dia, stats, objetivos, asignaciones, bloqueos_lunes, festivos, indisponibilidades
        )

        if pen1 == float("inf") or pen2 == float("inf"):
            continue

        # Penalizar si la pareja ya ha coincidido mucho
        coincidencias_previas = 0
        for personas_dia in asignaciones.values():
            if p1 in personas_dia and p2 in personas_dia:
                coincidencias_previas += 1

        pen_total = pen1 + pen2 + coincidencias_previas * 1.5

        if pen_total < mejor_pen:
            mejor_pen = pen_total
            mejor_pareja = (p1, p2)

    return mejor_pareja


def generar_planificacion(
    anio: int,
    personas: List[str],
    festivos: Set[date],
    indisponibilidades: Dict[str, Set[date]],
) -> Tuple[Dict[date, List[str]], Dict[str, EstadisticasPersona]]:
    """
    Genera la planificación anual respetando las reglas:
    - 2 personas por día
    - viernes -> mismo equipo también domingo
    - sábado -> lunes libre
    """
    fechas = generar_fechas_planificables(anio)
    stats = inicializar_estadisticas(personas)
    objetivos = calcular_objetivos_teoricos(fechas, len(personas), festivos, plazas_por_dia=2)
    asignaciones: Dict[date, List[str]] = {}

    fechas_set = set(fechas)

    for dia in fechas:
        t = tipo_dia(dia, festivos)

        # Recalcular bloqueos de lunes a partir de los sábados ya asignados
        bloqueos_lunes = crear_bloqueos_por_sabado(asignaciones)

        # Si es domingo y viene de un viernes planificable, debe repetir el mismo equipo
        if dia.weekday() == 6:
            viernes_prev = dia - timedelta(days=2)
            if viernes_prev in fechas_set and viernes_prev in asignaciones:
                pareja_viernes = asignaciones[viernes_prev]

                # Comprobación básica de viabilidad
                validos = True
                for persona in pareja_viernes:
                    if dia in indisponibilidades.get(persona, set()):
                        validos = False
                    if persona in bloqueos_lunes.get(dia, set()):
                        validos = False

                if validos:
                    asignaciones[dia] = pareja_viernes.copy()
                    for persona in pareja_viernes:
                        actualizar_estadisticas(stats, persona, dia, festivos)
                    continue

        # Para el resto de días, elegimos pareja óptima
        candidatos_ordenados = sorted(
            personas,
            key=lambda p: (
                stats[p].total,
                stats[p].festivos,
                stats[p].sabados + stats[p].domingos,
                stats[p].laborables,
            )
        )

        pareja = elegir_mejor_pareja(
            candidatos_ordenados,
            dia,
            stats,
            objetivos,
            asignaciones,
            bloqueos_lunes,
            festivos,
            indisponibilidades,
        )

        if pareja is None:
            # Relajación mínima: permitimos consecutividad solo si no hay otra solución
            pareja = elegir_mejor_pareja_relajada(
                candidatos_ordenados,
                dia,
                stats,
                objetivos,
                asignaciones,
                bloqueos_lunes,
                festivos,
                indisponibilidades,
            )

        if pareja is None:
            raise ValueError(
                f"No se ha podido asignar el día {dia}. "
                "Revisa indisponibilidades o endurecimiento de reglas."
            )

        asignaciones[dia] = [pareja[0], pareja[1]]
        actualizar_estadisticas(stats, pareja[0], dia, festivos)
        actualizar_estadisticas(stats, pareja[1], dia, festivos)

    return asignaciones, stats


def penalizacion_persona_relajada(
    persona: str,
    dia: date,
    stats: Dict[str, EstadisticasPersona],
    objetivos: Dict[str, float],
    asignaciones: Dict[date, List[str]],
    bloqueos_lunes: Dict[date, Set[str]],
    festivos: Set[date],
    indisponibilidades: Dict[str, Set[date]],
) -> float:
    """
    Versión relajada: permite días consecutivos si hace falta,
    pero lo penaliza mucho.
    """
    if dia in indisponibilidades.get(persona, set()):
        return float("inf")

    if persona in bloqueos_lunes.get(dia, set()):
        return float("inf")

    if persona in asignaciones.get(dia, []):
        return float("inf")

    st = stats[persona]
    t = tipo_dia(dia, festivos)
    pen = 0.0

    # Penalización fuerte por consecutividad
    dia_anterior = dia - timedelta(days=1)
    dia_siguiente = dia + timedelta(days=1)

    if persona in asignaciones.get(dia_anterior, []):
        pen += 50
    if persona in asignaciones.get(dia_siguiente, []):
        pen += 50

    pen += (st.total - objetivos["total"]) * 3

    if t == "laborable":
        pen += (st.laborables - objetivos["laborables"]) * 4
    elif t == "viernes":
        pen += (st.viernes - objetivos["viernes"]) * 5
    elif t == "sabado":
        pen += (st.sabados - objetivos["sabados"]) * 6
    elif t == "domingo":
        pen += (st.domingos - objetivos["domingos"]) * 6
    elif t == "festivo":
        pen += (st.festivos - objetivos["festivos"]) * 7

    pen += (st.sabados + st.domingos + st.festivos) * 0.8

    return pen


def elegir_mejor_pareja_relajada(
    candidatos: List[str],
    dia: date,
    stats: Dict[str, EstadisticasPersona],
    objetivos: Dict[str, float],
    asignaciones: Dict[date, List[str]],
    bloqueos_lunes: Dict[date, Set[str]],
    festivos: Set[date],
    indisponibilidades: Dict[str, Set[date]],
) -> Tuple[str, str] | None:
    mejor_pareja = None
    mejor_pen = float("inf")

    for p1, p2 in combinations(candidatos, 2):
        pen1 = penalizacion_persona_relajada(
            p1, dia, stats, objetivos, asignaciones, bloqueos_lunes, festivos, indisponibilidades
        )
        pen2 = penalizacion_persona_relajada(
            p2, dia, stats, objetivos, asignaciones, bloqueos_lunes, festivos, indisponibilidades
        )

        if pen1 == float("inf") or pen2 == float("inf"):
            continue

        coincidencias_previas = 0
        for personas_dia in asignaciones.values():
            if p1 in personas_dia and p2 in personas_dia:
                coincidencias_previas += 1

        pen_total = pen1 + pen2 + coincidencias_previas * 1.5

        if pen_total < mejor_pen:
            mejor_pen = pen_total
            mejor_pareja = (p1, p2)

    return mejor_pareja


# ============================================================
# EXPORTACIÓN
# ============================================================

def construir_dataframe_planificacion(
    asignaciones: Dict[date, List[str]],
    festivos: Set[date],
) -> pd.DataFrame:
    filas = []

    for dia in sorted(asignaciones.keys()):
        personas = asignaciones[dia]
        filas.append({
            "fecha": dia,
            "dia_semana": DIAS_ES[dia.weekday()],
            "tipo_dia": tipo_dia(dia, festivos),
            "persona_1": personas[0],
            "persona_2": personas[1],
            "es_festivo": "Sí" if dia in festivos else "No",
        })

    df = pd.DataFrame(filas)
    return df


def construir_dataframe_resumen(
    stats: Dict[str, EstadisticasPersona]
) -> pd.DataFrame:
    filas = []

    for persona, st in stats.items():
        filas.append({
            "persona": persona,
            "total_guardias": st.total,
            "laborables": st.laborables,
            "viernes": st.viernes,
            "sabados": st.sabados,
            "domingos": st.domingos,
            "festivos": st.festivos,
            "fin_de_semana_total": st.sabados + st.domingos,
        })

    df = pd.DataFrame(filas).sort_values(
        by=["total_guardias", "festivos", "fin_de_semana_total", "laborables", "persona"]
    ).reset_index(drop=True)

    return df


def exportar_excel(
    df_plan: pd.DataFrame,
    df_resumen: pd.DataFrame,
    ruta_excel: str = "planificacion_guardias_quirófano.xlsx",
) -> None:
    with pd.ExcelWriter(ruta_excel, engine="openpyxl") as writer:
        df_plan.to_excel(writer, sheet_name="Planificacion", index=False)
        df_resumen.to_excel(writer, sheet_name="Resumen", index=False)

        wb = writer.book
        ws_plan = writer.sheets["Planificacion"]
        ws_resumen = writer.sheets["Resumen"]

        # ============================================================
        # ESTILOS GENERALES
        # ============================================================
        color_cabecera = "1F4E78"
        color_texto_cabecera = "FFFFFF"

        color_laborable = "EAF2F8"
        color_viernes = "D4E6F1"
        color_sabado = "FCF3CF"
        color_domingo = "FADBD8"
        color_festivo = "F5CBA7"

        borde_fino = Border(
            left=Side(style="thin", color="D9D9D9"),
            right=Side(style="thin", color="D9D9D9"),
            top=Side(style="thin", color="D9D9D9"),
            bottom=Side(style="thin", color="D9D9D9"),
        )

        fuente_cabecera = Font(bold=True, color=color_texto_cabecera)
        relleno_cabecera = PatternFill("solid", fgColor=color_cabecera)

        alineacion_centrada = Alignment(horizontal="center", vertical="center")
        alineacion_izquierda = Alignment(horizontal="left", vertical="center")

        # ============================================================
        # FORMATO HOJA PLANIFICACION
        # ============================================================
        ws_plan.freeze_panes = "A2"
        ws_plan.auto_filter.ref = ws_plan.dimensions
        ws_plan.sheet_view.showGridLines = False

        # Cabeceras
        for cell in ws_plan[1]:
            cell.fill = relleno_cabecera
            cell.font = fuente_cabecera
            cell.alignment = alineacion_centrada
            cell.border = borde_fino

        # Filas de datos
        for row in ws_plan.iter_rows(min_row=2, max_row=ws_plan.max_row):
            tipo = row[2].value  # columna tipo_dia

            if tipo == "laborable":
                relleno = PatternFill("solid", fgColor=color_laborable)
            elif tipo == "viernes":
                relleno = PatternFill("solid", fgColor=color_viernes)
            elif tipo == "sabado":
                relleno = PatternFill("solid", fgColor=color_sabado)
            elif tipo == "domingo":
                relleno = PatternFill("solid", fgColor=color_domingo)
            elif tipo == "festivo":
                relleno = PatternFill("solid", fgColor=color_festivo)
            else:
                relleno = PatternFill("solid", fgColor="FFFFFF")

            for i, cell in enumerate(row, start=1):
                cell.fill = relleno
                cell.border = borde_fino

                if i == 1:
                    cell.number_format = "dd/mm/yyyy"
                    cell.alignment = alineacion_centrada
                elif i in [2, 3, 6]:
                    cell.alignment = alineacion_centrada
                else:
                    cell.alignment = alineacion_izquierda

        # Ajuste de ancho de columnas
        for col in ws_plan.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)

            for cell in col:
                try:
                    valor = str(cell.value) if cell.value is not None else ""
                    if len(valor) > max_length:
                        max_length = len(valor)
                except Exception:
                    pass

            ws_plan.column_dimensions[col_letter].width = max_length + 3

        # Altura de filas
        for row_num in range(1, ws_plan.max_row + 1):
            ws_plan.row_dimensions[row_num].height = 22

        # ============================================================
        # FORMATO HOJA RESUMEN
        # ============================================================
        ws_resumen.freeze_panes = "A2"
        ws_resumen.auto_filter.ref = ws_resumen.dimensions
        ws_resumen.sheet_view.showGridLines = False

        for cell in ws_resumen[1]:
            cell.fill = relleno_cabecera
            cell.font = fuente_cabecera
            cell.alignment = alineacion_centrada
            cell.border = borde_fino

        for row in ws_resumen.iter_rows(min_row=2, max_row=ws_resumen.max_row):
            for i, cell in enumerate(row, start=1):
                cell.border = borde_fino
                if i == 1:
                    cell.alignment = alineacion_izquierda
                else:
                    cell.alignment = alineacion_centrada

        # Ajuste de ancho de columnas resumen
        for col in ws_resumen.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)

            for cell in col:
                try:
                    valor = str(cell.value) if cell.value is not None else ""
                    if len(valor) > max_length:
                        max_length = len(valor)
                except Exception:
                    pass

            ws_resumen.column_dimensions[col_letter].width = max_length + 3

        for row_num in range(1, ws_resumen.max_row + 1):
            ws_resumen.row_dimensions[row_num].height = 22


# ============================================================
# MAIN
# ============================================================

if __name__ == "__main__":
    asignaciones, stats = generar_planificacion(
        anio=ANIO,
        personas=PERSONAS,
        festivos=FESTIVOS,
        indisponibilidades=INDISPONIBILIDADES,
    )

    df_plan = construir_dataframe_planificacion(asignaciones, FESTIVOS)
    df_resumen = construir_dataframe_resumen(stats)

    exportar_excel(df_plan, df_resumen, "planificacion_guardias_quirófano.xlsx")

    print("Planificación generada correctamente.")
    print()
    print("Resumen por persona:")
    print(df_resumen.to_string(index=False))
    print()
    print("Archivo Excel generado: planificacion_guardias_quirófano.xlsx")